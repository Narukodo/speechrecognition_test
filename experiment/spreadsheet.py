from collections import namedtuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EngineEntry = namedtuple('EngineEntry', ['engine_name', 'col_index'])


class Spreadsheet:
    def __init__(self):
        self.workbook = Workbook()
        self.filename = 'results.xlsx'
        self.tables = []
        self.next_col = 1

    def save(self):
        self.workbook.save('speech_engine_tests.xlsx')

    def get_table_column(self, speech_engine_name):
        for engine_entry in self.tables:
            if engine_entry.engine_name == speech_engine_name:
                return engine_entry.col_index
        raise Exception(f'table for {speech_engine_name} does not exist')

    def add_speech_engine(self, speech_engine_name):
        try:
            self.tables.index(speech_engine_name)
            raise Exception(f'{speech_engine_name} already exists')
        except:
            worksheet = self.workbook.active
            self.populate_headings(self.next_col, speech_engine_name, ['File Name', 'Word Error Rate'])
            self.next_col += 4
            self.save()
    
    def populate_headings(self, start_col, group_name, headings):
        worksheet = self.workbook.active
        worksheet.merge_cells(f'{get_column_letter(start_col)}1:{get_column_letter(start_col + len(headings) - 1)}1')
        worksheet[f'{get_column_letter(start_col)}1'] = group_name
        for idx, heading in enumerate(headings):
            worksheet[f'{get_column_letter(start_col + idx)}2'] = heading
        self.tables.append(EngineEntry(group_name, start_col))

    def populate_table(self, speech_engine_name, data):
        worksheet = self.workbook.active
        cur_col = self.get_table_column(speech_engine_name)
        file_name = get_column_letter(cur_col)
        wer = get_column_letter(cur_col + 1)
        row = 3
        for data_point in data:
            worksheet[f'{file_name}{row}'] = data_point.filename
            worksheet[f'{wer}{row}'] = data_point.wer
            row += 1
        self.save()
    